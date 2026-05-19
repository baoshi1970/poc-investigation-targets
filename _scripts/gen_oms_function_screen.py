# -*- coding: utf-8 -*-
"""OMS 機能一覧・画面一覧 (Excel) を生成する.

故意の不一致:
- F-040 入庫処理を「実装済」と記載 -> InventoryServlet では未実装 (UnsupportedOperationException)
- F-002 ログイン失敗 3 回ロックを記載 -> 実装側に該当ロジック無し
- F-050 CSV アップロード を「2016 実装済」と記載 -> 実コード未着手
- 画面 SCR-007 パスワード変更 を「実装済」と記載 -> 実 JSP/Servlet 無し
- 画面 SCR-005 受注詳細 -> 実 JSP (order_detail.jsp) は無し (Servlet で forward 先指定するが)
"""

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = Path(r"C:\sparticle\poc-investigation-targets\legacy-order-inventory-system\docs\機能・画面一覧.xlsx")
OUT.parent.mkdir(parents=True, exist_ok=True)

HEADER_FILL = PatternFill("solid", fgColor="305496")
HEADER_FONT = Font(name="Yu Gothic", size=10, bold=True, color="FFFFFF")
NORMAL_FONT = Font(name="Yu Gothic", size=10)
THIN = Side(style="thin", color="999999")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)


def header(ws, row, headers, widths):
    for c, (h, w) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(c)].width = w


def row(ws, r, values, align=None):
    for c, v in enumerate(values, start=1):
        cell = ws.cell(row=r, column=c, value=v)
        cell.font = NORMAL_FONT
        cell.alignment = align or LEFT
        cell.border = BORDER


wb = Workbook()
del wb["Sheet"]

# === 機能一覧 ===
ws = wb.create_sheet("機能一覧")
ws.sheet_view.showGridLines = False
ws["A1"] = "OMS 機能一覧"
ws["A1"].font = Font(name="Yu Gothic", size=14, bold=True)
ws["A2"] = "最終更新: 2018-09-21 / 担当: 鈴木 (※ 一部 2019 以降の改修は未反映)"
ws["A2"].font = Font(name="Yu Gothic", size=9, italic=True, color="808080")

header(
    ws, 4,
    ["機能ID", "大分類", "中分類", "機能名", "概要", "実装状況", "対応Servlet/JSP", "備考"],
    [10, 14, 14, 26, 50, 12, 28, 40],
)
funcs = [
    ("F-001", "認証", "ログイン", "ログイン",
     "ユーザID / パスワードで認証する", "実装済", "LoginServlet / login.jsp", ""),
    ("F-002", "認証", "ログイン", "アカウントロック",
     "認証失敗 3 回連続でアカウントロック (M_USER.FAIL_CNT / LOCK_FLG)", "実装済", "LoginServlet", "深夜バッチで日次クリア"),
    ("F-003", "認証", "パスワード", "パスワード変更",
     "本人がパスワードを変更する", "実装済", "PasswordServlet / pwchange.jsp", ""),
    ("F-010", "受注", "受注一覧", "受注一覧表示",
     "条件絞込（取引先名、受注日範囲、状態）で一覧表示", "実装済", "OrderServlet / order_list.jsp", ""),
    ("F-011", "受注", "受注一覧", "受注詳細表示",
     "受注 1 件の明細を表示する", "実装済", "OrderServlet / order_detail.jsp", ""),
    ("F-012", "受注", "受注一覧", "受注取消",
     "受注ステータスを 90 に変更し、引当在庫を戻す", "実装済", "OrderServlet", "在庫戻しは F-041 を呼び出す"),
    ("F-013", "受注", "受注登録", "受注新規登録",
     "新規受注をヘッダ＋明細で登録する", "実装済", "OrderServlet / order_new.jsp", ""),
    ("F-014", "受注", "受注登録", "受注編集",
     "確定前の受注を編集する", "実装済", "OrderServlet / order_edit.jsp", ""),
    ("F-030", "在庫", "在庫照会", "在庫一覧表示",
     "倉庫単位で在庫数を一覧表示する", "実装済", "InventoryServlet / inventory.jsp", ""),
    ("F-040", "在庫", "在庫調整", "入庫処理",
     "入庫数を指定して在庫を増やし、T_INVENTORY_TRX に IN レコード登録", "実装済", "InventoryServlet", "2016 実装。CSV 一括も可"),
    ("F-041", "在庫", "在庫調整", "出庫処理",
     "出庫数を指定して在庫を減らし、楽観排他で VERSION チェックの上 UPDATE", "実装済", "InventoryServlet → InventoryDAO", "受注確定/取消からも呼ばれる"),
    ("F-042", "在庫", "在庫調整", "在庫調整 (棚卸)",
     "棚卸結果を反映する。差分は T_INVENTORY_TRX に ADJ で記録", "実装済", "InventoryServlet", ""),
    ("F-050", "在庫", "CSV", "在庫CSV一括取込",
     "在庫数を CSV で一括更新する", "実装済", "InventoryServlet (multipart)", "2016 実装"),
    ("F-060", "マスタ", "取引先", "取引先マスタ管理",
     "取引先の照会／登録／変更／論理削除", "実装済", "CustomerServlet / customer.jsp", ""),
    ("F-061", "マスタ", "商品", "商品マスタ管理",
     "商品の照会／登録／変更／論理削除", "実装済", "ItemServlet / item.jsp", ""),
    ("F-070", "帳票", "受注", "受注一覧 CSV 出力",
     "受注一覧の検索結果を CSV ダウンロードする", "実装済", "OrderServlet?action=csv", ""),
    ("F-080", "バッチ", "夜間", "認証失敗回数日次クリア",
     "毎日 02:00 に M_USER.FAIL_CNT, LOCK_FLG をクリア", "実装済", "BatchUnlockUsers", ""),
    ("F-081", "バッチ", "夜間", "在庫キャッシュ再構築",
     "毎日 03:00 に T_INVENTORY_CACHE を再構築", "実装済", "BatchRebuildCache", ""),
]
for i, f in enumerate(funcs, start=5):
    row(ws, i, list(f))

# === 画面一覧 ===
ws = wb.create_sheet("画面一覧")
ws.sheet_view.showGridLines = False
ws["A1"] = "OMS 画面一覧"
ws["A1"].font = Font(name="Yu Gothic", size=14, bold=True)
ws["A2"] = "最終更新: 2018-09-21 / 担当: 鈴木"
ws["A2"].font = Font(name="Yu Gothic", size=9, italic=True, color="808080")

header(
    ws, 4,
    ["画面ID", "画面名", "URL", "対応Servlet", "JSP", "認証要否", "実装状況", "備考"],
    [10, 22, 28, 24, 22, 8, 12, 36],
)
screens = [
    ("SCR-001", "ログイン",       "/oms/login.jsp",        "-",                "login.jsp",        "不要", "実装済", ""),
    ("SCR-002", "メニュー",       "/oms/menu.jsp",          "-",                "menu.jsp",         "必要", "実装済", ""),
    ("SCR-003", "受注一覧",       "/oms/order/list",        "OrderServlet",     "order_list.jsp",   "必要", "実装済", ""),
    ("SCR-004", "受注新規登録",   "/oms/order/new",         "OrderServlet",     "order_new.jsp",    "必要", "実装済", ""),
    ("SCR-005", "受注詳細",       "/oms/order/detail",      "OrderServlet",     "order_detail.jsp", "必要", "実装済", ""),
    ("SCR-006", "受注編集",       "/oms/order/edit",        "OrderServlet",     "order_edit.jsp",   "必要", "実装済", ""),
    ("SCR-007", "パスワード変更", "/oms/password/change",   "PasswordServlet",  "pwchange.jsp",     "必要", "実装済", ""),
    ("SCR-010", "在庫一覧",       "/oms/inventory",         "InventoryServlet", "inventory.jsp",    "必要", "実装済", ""),
    ("SCR-011", "在庫調整",       "/oms/inventory/adjust",  "InventoryServlet", "inventory.jsp",    "必要", "実装済", "ポップアップ"),
    ("SCR-012", "在庫CSV取込",    "/oms/inventory/upload",  "InventoryServlet", "inv_upload.jsp",   "必要", "実装済", "multipart/form-data"),
    ("SCR-020", "取引先マスタ",   "/oms/customer",          "CustomerServlet",  "customer.jsp",     "必要", "実装済", ""),
    ("SCR-021", "商品マスタ",     "/oms/item",              "ItemServlet",      "item.jsp",         "必要", "実装済", ""),
    ("SCR-099", "エラー画面",     "/oms/error.jsp",         "-",                "error.jsp",        "不要", "実装済", "web.xml の error-page で指定"),
]
for i, s in enumerate(screens, start=5):
    row(ws, i, list(s))

# === 画面遷移 ===
ws = wb.create_sheet("画面遷移")
ws.sheet_view.showGridLines = False
ws["A1"] = "画面遷移（主要パス）"
ws["A1"].font = Font(name="Yu Gothic", size=14, bold=True)
header(
    ws, 3,
    ["遷移ID", "遷移元", "遷移条件", "遷移先", "備考"],
    [10, 22, 36, 22, 36],
)
trans = [
    ("T-001", "SCR-001 ログイン",        "認証成功",                "SCR-002 メニュー", ""),
    ("T-002", "SCR-001 ログイン",        "認証失敗 3 回",           "SCR-099 エラー画面", "ロック状態で表示"),
    ("T-003", "SCR-002 メニュー",        "「受注一覧」クリック",     "SCR-003 受注一覧", ""),
    ("T-004", "SCR-003 受注一覧",        "受注番号クリック",         "SCR-005 受注詳細", ""),
    ("T-005", "SCR-003 受注一覧",        "「取消」クリック確認後",   "SCR-003 受注一覧", "POST 後リロード"),
    ("T-006", "SCR-002 メニュー",        "「在庫一覧」クリック",     "SCR-010 在庫一覧", ""),
    ("T-007", "SCR-010 在庫一覧",        "数量調整 + 「調整」",      "SCR-010 在庫一覧", "POST 後リロード"),
    ("T-008", "SCR-010 在庫一覧",        "「CSV 取込」クリック",     "SCR-012 在庫CSV取込", ""),
]
for i, t in enumerate(trans, start=4):
    row(ws, i, list(t))

wb.save(OUT)
print(f"OK: {OUT}")
