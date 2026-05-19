# -*- coding: utf-8 -*-
"""Asteria 連携先 DWH テーブル定義書 (Excel).

故意の不一致:
- DIM_INVENTORY_W003 を「W003 専用テーブル」として記載 -> 実 ETL は DIM_INVENTORY を見ている (FLW-003 不具合の遠因)
- FACT_ORDER に「PARTITION BY ORDER_DT」と記載 -> 実 BULK INSERT は TRUNCATE → 全件のため意味なし
- DIM_CUSTOMER は「MERGE 方式更新」と記載 -> 実装は TRUNCATE → INSERT
- TAX_RATE の値域に「0.10 / 0.08 / 0.00」と記載 -> 軽減税率/標準税率改定後の値が未反映
"""

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = Path(r"C:\sparticle\poc-investigation-targets\asteria-etl-sample\docs\DWH連携テーブル定義書.xlsx")
OUT.parent.mkdir(parents=True, exist_ok=True)

HEADER_FILL = PatternFill("solid", fgColor="305496")
HEADER_FONT = Font(name="Yu Gothic", size=10, bold=True, color="FFFFFF")
NORMAL_FONT = Font(name="Yu Gothic", size=10)
PK_FONT     = Font(name="Yu Gothic", size=10, bold=True, color="C00000")
THIN = Side(style="thin", color="999999")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)


def header(ws, row, headers, widths):
    for c, (h, w) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(c)].width = w


def row(ws, r, values, pk=False, align=None):
    for c, v in enumerate(values, start=1):
        cell = ws.cell(row=r, column=c, value=v)
        cell.font = PK_FONT if pk else NORMAL_FONT
        cell.alignment = align or LEFT
        cell.border = BORDER


def make_sheet(wb, name, meta, cols):
    ws = wb.create_sheet(name)
    ws.sheet_view.showGridLines = False
    ws["A1"] = "DWH テーブル定義書"
    ws["A1"].font = Font(name="Yu Gothic", size=14, bold=True)

    r = 3
    for k, v in meta:
        ws.cell(row=r, column=1, value=k).font = Font(name="Yu Gothic", size=10, bold=True)
        ws.cell(row=r, column=1).fill = PatternFill("solid", fgColor="DDEBF7")
        ws.cell(row=r, column=1).border = BORDER
        ws.cell(row=r, column=1).alignment = LEFT
        c2 = ws.cell(row=r, column=2, value=v)
        c2.font = NORMAL_FONT
        c2.border = BORDER
        c2.alignment = LEFT
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
        r += 1
    ws.column_dimensions["A"].width = 20
    for col in "BCDEF":
        ws.column_dimensions[col].width = 14

    r += 1
    ws.cell(row=r, column=1, value="■ カラム").font = Font(name="Yu Gothic", size=11, bold=True)
    r += 1
    header(ws, r, ["No", "物理名", "論理名", "型", "必須", "備考"], [4, 22, 24, 16, 6, 50])
    r += 1
    for i, col in enumerate(cols, start=1):
        row(ws, r,
            [i, col["phys"], col["logical"], col["type"], "○" if col["nn"] else "", col.get("note", "")],
            pk=col.get("pk", False))
        r += 1


wb = Workbook()
del wb["Sheet"]

# === 改訂履歴 ===
ws = wb.create_sheet("改訂履歴")
ws.sheet_view.showGridLines = False
ws["A1"] = "DWH 連携テーブル定義書 - 改訂履歴"
ws["A1"].font = Font(name="Yu Gothic", size=14, bold=True)
header(ws, 3, ["版", "改訂日", "改訂者", "改訂内容"], [6, 14, 14, 90])
hist = [
    ("1.0", "2017-06-15", "外部SI",  "初版 (Asteria Warp 1709 導入)"),
    ("1.1", "2019-10-01", "鈴木",    "軽減税率対応で TAX_RATE 追加"),
    ("1.2", "2022-04-20", "鈴木",    "W003 倉庫追加に伴い DIM_INVENTORY_W003 新設 (※ Asteria 側未対応)"),
    ("1.3", "2024-03-15", "鈴木",    "税率改定対応 (DWH 取込側のみ反映、Asteria の XSLT 側未対応)"),
]
for i, h in enumerate(hist, start=4):
    row(ws, i, list(h))

# === DIM_CUSTOMER ===
make_sheet(
    wb, "DIM_CUSTOMER",
    [
        ("テーブル物理名", "DIM_CUSTOMER"),
        ("テーブル論理名", "取引先ディメンション"),
        ("用途", "取引先マスタの DWH 側コピー"),
        ("更新方式", "MERGE 方式 (キー一致は UPDATE、新規は INSERT、未連携は DEL_FLG='1')"),
        ("更新タイミング", "Asteria フロー FLW-001 customer_master_sync 経由、平日 06:00"),
        ("想定件数", "約 8 万件"),
    ],
    [
        {"phys": "CUST_CODE",   "logical": "取引先コード", "type": "VARCHAR(10)",  "nn": True, "pk": True, "note": "主キー"},
        {"phys": "CUST_NAME",   "logical": "取引先名",     "type": "NVARCHAR(120)","nn": True, "note": "Unicode 対応 (半角カナ・機種依存文字も可)"},
        {"phys": "PHONE",       "logical": "電話番号",     "type": "VARCHAR(20)",  "nn": False, "note": ""},
        {"phys": "ADDRESS",     "logical": "住所",         "type": "NVARCHAR(200)","nn": False, "note": ""},
        {"phys": "DEL_FLG",     "logical": "削除フラグ",   "type": "CHAR(1)",      "nn": True,  "note": "0:有効 / 1:削除"},
        {"phys": "LOAD_AT",     "logical": "取込日時",     "type": "DATETIME2",    "nn": True,  "note": "Asteria 取込時の日時"},
    ],
)

# === DIM_INVENTORY ===
make_sheet(
    wb, "DIM_INVENTORY",
    [
        ("テーブル物理名", "DIM_INVENTORY"),
        ("テーブル論理名", "在庫ディメンション"),
        ("用途", "在庫マスタの DWH 側コピー (W001, W002 のみ格納)"),
        ("更新方式", "差分 MERGE"),
        ("更新タイミング", "Asteria フロー FLW-003 inventory_reconcile 経由、月初 04:00"),
        ("想定件数", "約 5 万件 (W001 + W002)"),
    ],
    [
        {"phys": "ITEM_CODE",   "logical": "商品コード", "type": "VARCHAR(13)", "nn": True, "pk": True, "note": ""},
        {"phys": "WAREHOUSE",   "logical": "倉庫コード", "type": "VARCHAR(4)",  "nn": True, "pk": True, "note": "W001 / W002 (W003 は別テーブル)"},
        {"phys": "STOCK_QTY",   "logical": "在庫数",     "type": "INT",         "nn": True, "note": ""},
        {"phys": "LOAD_AT",     "logical": "取込日時",   "type": "DATETIME2",   "nn": True, "note": ""},
    ],
)

# === DIM_INVENTORY_W003 ===
make_sheet(
    wb, "DIM_INVENTORY_W003",
    [
        ("テーブル物理名", "DIM_INVENTORY_W003"),
        ("テーブル論理名", "在庫ディメンション (W003 福岡倉庫専用)"),
        ("用途", "W003 倉庫の在庫を別管理する (2022 追加。倉庫運用主体が異なるため別テーブル化)"),
        ("更新方式", "差分 MERGE"),
        ("更新タイミング", "Asteria フロー FLW-003 経由 (W003 用に分岐実装される予定)"),
        ("想定件数", "約 1.5 万件"),
    ],
    [
        {"phys": "ITEM_CODE",   "logical": "商品コード", "type": "VARCHAR(13)", "nn": True, "pk": True, "note": ""},
        {"phys": "STOCK_QTY",   "logical": "在庫数",     "type": "INT",         "nn": True, "note": ""},
        {"phys": "LOAD_AT",     "logical": "取込日時",   "type": "DATETIME2",   "nn": True, "note": ""},
    ],
)

# === FACT_ORDER ===
make_sheet(
    wb, "FACT_ORDER",
    [
        ("テーブル物理名", "FACT_ORDER"),
        ("テーブル論理名", "受注ファクト"),
        ("用途", "受注ヘッダの DWH 取込先"),
        ("更新方式", "差分追記 (新規/更新のみ MERGE、削除は論理削除フラグで管理)"),
        ("パーティション", "ORDER_DT 月単位パーティション (RANGE 分割)"),
        ("更新タイミング", "Asteria フロー FLW-002 order_etl_daily 経由、毎日 02:30"),
        ("想定件数", "月 50 万件 / 累計 3,000 万件以上"),
    ],
    [
        {"phys": "ORDER_NO",     "logical": "受注番号",  "type": "VARCHAR(14)",  "nn": True, "pk": True, "note": ""},
        {"phys": "ORDER_DT",     "logical": "受注日",    "type": "DATE",         "nn": True, "note": "パーティションキー"},
        {"phys": "CUSTOMER_CD",  "logical": "取引先コード","type": "VARCHAR(10)","nn": True, "note": "DIM_CUSTOMER と結合"},
        {"phys": "CUSTOMER_NM",  "logical": "取引先名",  "type": "NVARCHAR(120)","nn": True, "note": "正規化のため非推奨だが性能のため保持"},
        {"phys": "STATUS_CD",    "logical": "ステータス","type": "VARCHAR(2)",   "nn": True, "note": "00/10/90"},
        {"phys": "TOTAL_AMT",    "logical": "合計金額(税抜)","type": "DECIMAL(14,2)","nn": True, "note": ""},
        {"phys": "TAX_RATE",     "logical": "税率",      "type": "DECIMAL(5,3)", "nn": True, "note": "0.10 / 0.08 / 0.00 のいずれか"},
        {"phys": "TAX_EXCL_AMT", "logical": "税抜金額",  "type": "DECIMAL(14,2)","nn": True, "note": "TOTAL_AMT / (1 + TAX_RATE)"},
        {"phys": "SALES_CHANNEL","logical": "販売チャネル","type": "VARCHAR(10)","nn": True, "note": "STORE / EC / WHOLESALE"},
        {"phys": "CURRENCY",     "logical": "通貨",      "type": "VARCHAR(3)",   "nn": True, "note": "JPY 固定"},
        {"phys": "LOAD_AT",      "logical": "取込日時",  "type": "DATETIME2",    "nn": True, "note": ""},
    ],
)

wb.save(OUT)
print("OK")
