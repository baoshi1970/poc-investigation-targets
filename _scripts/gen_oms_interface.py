# -*- coding: utf-8 -*-
"""OMS I/F 仕様書 (Excel) を生成する.

故意の不一致:
- 旧 EDI 連携 (IF-001) を「稼働中」と表記 -> 実コードに該当無し (廃止済の残骸)
- 連携先システム名が古いまま (旧 SCM が現在は別ベンダー製に置き換わっている)
- IF-005 (DWH 連携) は実は ASTERIA フロー側で送信されているが本書では「OMS 側 OrderExportBatch から送信」と誤記
- IF-010 配送センター連携の文字コードを UTF-8 と記載 -> 実は CP932 で長年運用中
"""

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = Path(r"C:\sparticle\poc-investigation-targets\legacy-order-inventory-system\docs\IF仕様書.xlsx")
OUT.parent.mkdir(parents=True, exist_ok=True)

HEADER_FILL = PatternFill("solid", fgColor="305496")
HEADER_FONT = Font(name="Yu Gothic", size=10, bold=True, color="FFFFFF")
NORMAL_FONT = Font(name="Yu Gothic", size=10)
THIN = Side(style="thin", color="999999")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)


def header(ws, row, headers, widths):
    for c, (h, w) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(c)].width = w


def row(ws, r, values, align=None):
    for c, v in enumerate(values, start=1):
        cell = ws.cell(row=r, column=c, value=v)
        cell.font = NORMAL_FONT
        cell.alignment = align or LEFT
        cell.border = BORDER


wb = Workbook()
del wb["Sheet"]

# === 連携一覧 ===
ws = wb.create_sheet("連携一覧")
ws.sheet_view.showGridLines = False
ws["A1"] = "OMS 外部連携 I/F 一覧"
ws["A1"].font = Font(name="Yu Gothic", size=14, bold=True)
ws["A2"] = "最終更新: 2016-04-10 / 担当: 佐藤  (※ 2018 以降の追加・廃止は未反映)"
ws["A2"].font = Font(name="Yu Gothic", size=9, italic=True, color="808080")

header(
    ws, 4,
    ["I/F ID", "I/F 名", "方向", "連携先", "プロトコル", "形式", "文字コード", "頻度", "稼働状況", "備考"],
    [9, 28, 7, 22, 14, 12, 11, 14, 10, 36],
)
ifs = [
    ("IF-001", "旧 EDI 受注取込",         "受信", "EDI センター (旧 SCM 社)",  "FTP",         "固定長",   "EBCDIC",    "随時",        "稼働中", "JOB_EDI_RECV → OMS 取込"),
    ("IF-002", "取引先マスタ受信",         "受信", "基幹 (販管) システム",       "FTP",         "CSV",      "Shift_JIS", "日次 01:30", "稼働中", "M_CUSTOMER 全件洗い替え"),
    ("IF-003", "商品マスタ受信",           "受信", "商品管理システム (MDM)",     "HTTP POST",   "XML",      "UTF-8",     "日次 02:00", "稼働中", "M_ITEM 差分更新"),
    ("IF-004", "受注明細出力 (会計)",       "送信", "会計システム",               "ファイル共有", "CSV",     "Shift_JIS", "日次 23:00", "稼働中", "JOB_ORDER_FOR_KEIRI"),
    ("IF-005", "受注明細出力 (DWH)",        "送信", "DWH",                       "ファイル共有", "TSV",     "UTF-8",     "日次 02:30", "稼働中", "OrderExportBatch から出力"),
    ("IF-006", "在庫スナップショット出力",   "送信", "経営ダッシュボード",         "ファイル共有", "CSV",     "UTF-8",     "日次 04:30", "稼働中", "前日終端時点の在庫"),
    ("IF-010", "配送指示連携",             "送信", "配送センター WMS",           "SFTP",        "固定長",   "UTF-8",     "随時 (受注確定即時)", "稼働中", "WMS_INSTRUCT.dat"),
    ("IF-020", "売上実績連携",             "送信", "本社経理 BI",                "DB Link",     "テーブル",  "-",         "日次 05:00", "稼働中", "BI 側 stg_oms_sales テーブル"),
    ("IF-099", "障害監視通知",             "送信", "監視サーバ Zabbix",          "SMTP",        "メール",   "UTF-8",     "障害時",     "稼働中", "OmsErrorNotifier 経由"),
]
for i, f in enumerate(ifs, start=5):
    row(ws, i, list(f))

# === IF-005 詳細 ===
ws = wb.create_sheet("IF-005_詳細")
ws.sheet_view.showGridLines = False
ws["A1"] = "I/F 仕様書 - IF-005 受注明細出力 (DWH)"
ws["A1"].font = Font(name="Yu Gothic", size=14, bold=True)

ws["A3"] = "■ 基本情報"
ws["A3"].font = Font(name="Yu Gothic", size=11, bold=True)
meta = [
    ("I/F ID", "IF-005"),
    ("I/F 名", "受注明細出力 (DWH)"),
    ("方向", "OMS → DWH (送信)"),
    ("起動方法", "OMS バッチ OrderExportBatch (日次 02:30 起動)"),
    ("ファイル配置先", r"\\fileserver01\dwh_stage\order\order_YYYYMMDDHHMM.tsv"),
    ("形式", "TSV (タブ区切り、ヘッダ行あり)"),
    ("文字コード", "UTF-8 (BOM 無し)"),
    ("改行コード", "LF"),
    ("命名規則", "order_YYYYMMDDHHMM.tsv (HHMM は OMS バッチ起動時刻)"),
    ("世代管理", "DWH 側で取込後に削除 (OMS 側からは削除しない)"),
]
for i, (k, v) in enumerate(meta, start=4):
    ws.cell(row=i, column=1, value=k).font = Font(name="Yu Gothic", size=10, bold=True)
    ws.cell(row=i, column=1).fill = PatternFill("solid", fgColor="DDEBF7")
    ws.cell(row=i, column=1).border = BORDER
    ws.cell(row=i, column=1).alignment = LEFT
    ws.cell(row=i, column=2, value=v).font = NORMAL_FONT
    ws.cell(row=i, column=2).border = BORDER
    ws.cell(row=i, column=2).alignment = LEFT
    ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=8)
ws.column_dimensions["A"].width = 18
for col in "BCDEFGH":
    ws.column_dimensions[col].width = 12

r = 4 + len(meta) + 1
ws.cell(row=r, column=1, value="■ レコード仕様").font = Font(name="Yu Gothic", size=11, bold=True)
r += 1
header(
    ws, r,
    ["No", "項目名", "物理名", "型", "桁", "必須", "備考"],
    [4, 22, 18, 10, 6, 6, 50],
)
r += 1
fields = [
    (1,  "受注番号",      "ORDER_NO",     "文字列", 14, "○", "OMS T_ORDER.ORDER_NO"),
    (2,  "受注日",        "ORDER_DT",     "日付",   10, "○", "YYYY-MM-DD"),
    (3,  "取引先コード",  "CUSTOMER_CD",  "文字列", 10, "○", ""),
    (4,  "取引先名",      "CUSTOMER_NM",  "文字列",120, "○", ""),
    (5,  "ステータス",    "STATUS_CD",    "文字列",  2, "○", "00/10/90"),
    (6,  "合計金額",      "TOTAL_AMT",    "数値",   14, "○", "税抜（DWH 側で再計算しないよう税抜で渡す）"),
    (7,  "税率",          "TAX_RATE",     "数値",    5, "○", "0.10 等"),
    (8,  "販売チャネル",  "SALES_CHANNEL","文字列", 10, "○", "STORE/EC/WHOLESALE"),
    (9,  "通貨",          "CURRENCY",     "文字列",  3, "○", "JPY 固定"),
    (10, "作成日時",      "CRT_DT",       "日時",   19, "○", "YYYY-MM-DD HH:MM:SS"),
    (11, "更新日時",      "UPD_DT",       "日時",   19, "○", "YYYY-MM-DD HH:MM:SS"),
]
for f in fields:
    row(ws, r, list(f))
    r += 1

r += 1
ws.cell(row=r, column=1, value="■ 異常系").font = Font(name="Yu Gothic", size=11, bold=True)
r += 1
err = [
    ("E-01", "ファイル出力失敗",       "OmsErrorNotifier 経由でメール通知。DWH 側は前日ファイルで運用継続。"),
    ("E-02", "DWH 側取込失敗",         "DWH 側からメール連絡。OMS 側はファイル再生成 (バッチ再実行) で対応。"),
    ("E-03", "ファイル名重複",         "OMS バッチ起動時刻 (HHMM) を変えて再実行する。"),
]
header(ws, r, ["No", "ケース", "対応"], [6, 30, 60])
r += 1
for e in err:
    row(ws, r, list(e))
    r += 1

# === IF-010 詳細 ===
ws = wb.create_sheet("IF-010_詳細")
ws.sheet_view.showGridLines = False
ws["A1"] = "I/F 仕様書 - IF-010 配送指示連携"
ws["A1"].font = Font(name="Yu Gothic", size=14, bold=True)

ws["A3"] = "■ 基本情報"
ws["A3"].font = Font(name="Yu Gothic", size=11, bold=True)
meta = [
    ("I/F ID", "IF-010"),
    ("I/F 名", "配送指示連携"),
    ("方向", "OMS → 配送センター WMS"),
    ("プロトコル", "SFTP"),
    ("形式", "固定長レコード"),
    ("文字コード", "UTF-8"),
    ("頻度", "随時 (受注確定タイミングで即時送信)"),
    ("ファイル名", "WMS_INSTRUCT_YYYYMMDDHHMMSS.dat"),
]
for i, (k, v) in enumerate(meta, start=4):
    ws.cell(row=i, column=1, value=k).font = Font(name="Yu Gothic", size=10, bold=True)
    ws.cell(row=i, column=1).fill = PatternFill("solid", fgColor="DDEBF7")
    ws.cell(row=i, column=1).border = BORDER
    ws.cell(row=i, column=1).alignment = LEFT
    ws.cell(row=i, column=2, value=v).font = NORMAL_FONT
    ws.cell(row=i, column=2).border = BORDER
    ws.cell(row=i, column=2).alignment = LEFT
    ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=6)
ws.column_dimensions["A"].width = 16
for col in "BCDEF":
    ws.column_dimensions[col].width = 14

r = 4 + len(meta) + 1
ws.cell(row=r, column=1, value="■ 固定長レイアウト").font = Font(name="Yu Gothic", size=11, bold=True)
r += 1
header(ws, r, ["項目No", "項目名", "開始位置", "バイト数", "型", "備考"], [6, 22, 8, 8, 10, 40])
r += 1
layout = [
    (1,  "レコード区分",       1,   1, "X",   "1: 明細"),
    (2,  "受注番号",           2,  14, "X",   ""),
    (3,  "明細番号",          16,   4, "9",   "ゼロパディング"),
    (4,  "商品コード",        20,  13, "X",   "JAN"),
    (5,  "数量",              33,   8, "9",   "ゼロパディング"),
    (6,  "倉庫コード",        41,   4, "X",   "W001 等"),
    (7,  "出荷指定日",        45,   8, "9",   "YYYYMMDD"),
    (8,  "予備",              53,  47, "X",   "スペース埋め"),
]
for f in layout:
    row(ws, r, list(f))
    r += 1

wb.save(OUT)
print("OK")
