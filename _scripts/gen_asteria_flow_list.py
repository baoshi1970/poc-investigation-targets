# -*- coding: utf-8 -*-
"""Asteria フロー一覧 (Excel).

故意の不一致:
- FLW-099 legacy_csv_export を「現役 (休止中)」と記載 -> project.xml では enabled=false (廃止候補)
- 各フローの「想定所要時間」が現実と大きく乖離
- FLW-002 を「差分更新方式」と記載 -> 実装は全件 TRUNCATE
- 異常時のリトライ 3 回と記載 -> 実装は無し
"""

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = Path(r"C:\sparticle\poc-investigation-targets\asteria-etl-sample\docs\フロー一覧.xlsx")
OUT.parent.mkdir(parents=True, exist_ok=True)

HEADER_FILL = PatternFill("solid", fgColor="305496")
HEADER_FONT = Font(name="Yu Gothic", size=10, bold=True, color="FFFFFF")
NORMAL_FONT = Font(name="Yu Gothic", size=10)
THIN = Side(style="thin", color="999999")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)


def header(ws, row, headers, widths):
    for c, (h, w) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(c)].width = w


def row(ws, r, values, align=None):
    for c, v in enumerate(values, start=1):
        cell = ws.cell(row=r, column=c, value=v)
        cell.font = NORMAL_FONT
        cell.alignment = align or LEFT
        cell.border = BORDER


wb = Workbook()
del wb["Sheet"]

# === フロー一覧 ===
ws = wb.create_sheet("フロー一覧")
ws.sheet_view.showGridLines = False
ws["A1"] = "ASTERIA Warp フロー一覧"
ws["A1"].font = Font(name="Yu Gothic", size=14, bold=True)
ws["A2"] = "最終更新: 2022-05-02 / 担当: 鈴木  (※ 以降の運用上の事象は別途台帳管理)"
ws["A2"].font = Font(name="Yu Gothic", size=9, italic=True, color="808080")

header(
    ws, 4,
    ["フローID", "フロー名", "用途", "起動方式", "実行スケジュール", "想定所要", "更新方式", "状態", "備考"],
    [10, 24, 30, 14, 18, 10, 16, 10, 30],
)
flows = [
    ("FLW-001", "customer_master_sync",   "取引先マスタを OMS → DWH へ連携",
     "Cron",       "平日 06:00",           "8 分",   "MERGE",          "現役",       ""),
    ("FLW-002", "order_etl_daily",        "受注ヘッダを OMS → DWH へ連携 (前日差分)",
     "Cron",       "毎日 02:30",           "45 分",  "差分更新 (MERGE)", "現役",       ""),
    ("FLW-003", "inventory_reconcile",    "OMS と DWH の在庫数突合 (差分レポート出力)",
     "Cron",       "月初 04:00",           "90 分",  "差分抽出 (Merger)", "現役",      "W003 は別フロー予定"),
    ("FLW-099", "legacy_csv_export",      "旧 BI 向け CSV 出力 (現 BI 移行後も保険として保持)",
     "手動",       "随時 (手動起動)",     "30 分",  "全件",            "現役 (休止中)", "完全廃止は要確認"),
]
for i, f in enumerate(flows, start=5):
    row(ws, i, list(f))

# === 異常系・連絡先 ===
ws = wb.create_sheet("運用ルール")
ws.sheet_view.showGridLines = False
ws["A1"] = "運用ルール"
ws["A1"].font = Font(name="Yu Gothic", size=14, bold=True)

rules = [
    ("通知先",          "it-ops@example.co.jp (フロー完了 / 失敗時)"),
    ("失敗時の挙動",    "Asteria 内蔵スケジューラのリトライ機能で 3 回までリトライ。3 回失敗で停止し SMTP 通知。"),
    ("障害連絡フロー",  "1. it-ops に通知 → 2. 鈴木 (兼任) が当日中に確認 → 3. 必要に応じて手動再実行"),
    ("手動再実行手順",  "Warp Designer から該当フローを開き、右クリック「実行」。引数は不要。"),
    ("実行ログ保管",    "Warp ログサーバ \\warpsrv01\\logs\\flow に 90 日保管。それ以降は自動削除。"),
    ("メンテ時間",      "毎月第 1 日曜 23:00-翌 02:00 はメンテのためフロー自動停止。"),
]
r = 3
for k, v in rules:
    ws.cell(row=r, column=1, value=k).font = Font(name="Yu Gothic", size=10, bold=True)
    ws.cell(row=r, column=1).fill = PatternFill("solid", fgColor="DDEBF7")
    ws.cell(row=r, column=1).border = BORDER
    ws.cell(row=r, column=1).alignment = LEFT
    c2 = ws.cell(row=r, column=2, value=v)
    c2.font = NORMAL_FONT
    c2.border = BORDER
    c2.alignment = LEFT
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    r += 1
ws.column_dimensions["A"].width = 18
for col in "BCDEF":
    ws.column_dimensions[col].width = 18

# === コンポーネント一覧 (FLW-002) ===
ws = wb.create_sheet("FLW-002コンポーネント")
ws.sheet_view.showGridLines = False
ws["A1"] = "FLW-002 order_etl_daily コンポーネント一覧"
ws["A1"].font = Font(name="Yu Gothic", size=14, bold=True)
header(
    ws, 3,
    ["No", "コンポーネントID", "種類", "名称", "概要"],
    [4, 12, 14, 24, 56],
)
comps = [
    (1, "O1", "DBInput",   "受注ヘッダ前日分取得", "T_ORDER から ORDER_DT = 前日 のレコードを抽出 (差分更新)"),
    (2, "O2", "LookupDB",  "取引先名引き当て",     "M_CUSTOMER とキャッシュありで Lookup"),
    (3, "O3", "XSLT",      "XSLT で変換",          "order_transform.xsl により DWH 形式へ変換 (税率計算含む)"),
    (4, "O4", "FileOutput","DWH 取込用 TSV 出力",  "ファイル共有経由で出力"),
    (5, "O5", "DBExecute", "DWH MERGE 実行",       "FACT_ORDER に対し受注番号キーで MERGE"),
    (6, "O6", "MailSend",  "完了通知",             "SMTP 通知 (件名は処理結果に応じて [OK]/[FAILED] 切替)"),
]
for i, c in enumerate(comps, start=4):
    row(ws, i, list(c))

wb.save(OUT)
print("OK")
