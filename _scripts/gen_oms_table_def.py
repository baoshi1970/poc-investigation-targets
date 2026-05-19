# -*- coding: utf-8 -*-
"""OMS テーブル定義書 (Excel) を生成する.

故意に実コードと食い違うポイント:
- T_INVENTORY に VERSION 列 (楽観排他用) と記載 -> 実テーブルに無い (ISSUE-178 の根本原因の隠れポイント)
- M_USER.PASSWORD の暗号方式を SHA-256 と記載 -> 実装は MD5
- 文字コードを UTF-8 と記載 -> 実装は Windows-31J
- 命名規約に「全テーブル T_xxx / M_xxx」と記載 -> 実装は InventoryTrx (キャメル+小文字) 存在
- T_INVENTORY_CACHE を「使用中」と記載 -> 実テーブルは無し (rebuildAllCache の dead path)
- インデックス一覧に「IDX_ORDER_DT」等を記載 -> 実 schema.sql には無し
"""

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = Path(r"C:\sparticle\poc-investigation-targets\legacy-order-inventory-system\docs\テーブル定義書.xlsx")
OUT.parent.mkdir(parents=True, exist_ok=True)

HEADER_FILL = PatternFill("solid", fgColor="305496")
HEADER_FONT = Font(name="Yu Gothic", size=10, bold=True, color="FFFFFF")
NORMAL_FONT = Font(name="Yu Gothic", size=10)
PK_FONT     = Font(name="Yu Gothic", size=10, bold=True, color="C00000")
THIN = Side(style="thin", color="999999")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)


def apply_header(ws, row, headers, widths):
    for c, (h, w) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(c)].width = w


def write_row(ws, row, values, is_pk=False):
    for c, v in enumerate(values, start=1):
        cell = ws.cell(row=row, column=c, value=v)
        cell.font = PK_FONT if is_pk else NORMAL_FONT
        cell.alignment = LEFT
        cell.border = BORDER


def make_table_sheet(wb, sheet_name, table_meta, columns, indexes):
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False

    # ヘッダ部
    ws["A1"] = "テーブル定義書"
    ws["A1"].font = Font(name="Yu Gothic", size=14, bold=True)
    ws.merge_cells("A1:F1")

    meta_pairs = [
        ("テーブル物理名", table_meta["phys"]),
        ("テーブル論理名", table_meta["logical"]),
        ("用途",           table_meta["purpose"]),
        ("作成日",         table_meta["created"]),
        ("最終更新日",     table_meta["updated"]),
        ("最終更新者",     table_meta["author"]),
        ("文字コード",     "UTF-8"),
        ("テーブル空間",   table_meta.get("tablespace", "OMS_DATA")),
    ]
    r = 3
    for k, v in meta_pairs:
        ws.cell(row=r, column=1, value=k).font = Font(name="Yu Gothic", size=10, bold=True)
        ws.cell(row=r, column=1).fill = PatternFill("solid", fgColor="DDEBF7")
        ws.cell(row=r, column=1).border = BORDER
        ws.cell(row=r, column=1).alignment = LEFT
        c2 = ws.cell(row=r, column=2, value=v)
        c2.font = NORMAL_FONT
        c2.alignment = LEFT
        c2.border = BORDER
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="■ カラム定義").font = Font(name="Yu Gothic", size=11, bold=True)
    r += 1
    apply_header(
        ws, r,
        ["No", "物理名", "論理名", "型/桁", "必須", "備考"],
        [5, 22, 26, 18, 6, 50],
    )
    r += 1
    for i, col in enumerate(columns, start=1):
        write_row(ws, r, [i, col["phys"], col["logical"], col["type"],
                          "○" if col["nn"] else "", col.get("note", "")],
                  is_pk=col.get("pk", False))
        r += 1

    if indexes:
        r += 1
        ws.cell(row=r, column=1, value="■ インデックス").font = Font(name="Yu Gothic", size=11, bold=True)
        r += 1
        apply_header(ws, r, ["No", "索引名", "種類", "対象カラム", "備考", ""],
                     [5, 22, 12, 30, 35, 5])
        r += 1
        for i, idx in enumerate(indexes, start=1):
            write_row(ws, r, [i, idx["name"], idx["kind"], idx["cols"], idx.get("note", ""), ""])
            r += 1


wb = Workbook()
ws0 = wb.active
ws0.title = "改訂履歴"
ws0.sheet_view.showGridLines = False
ws0["A1"] = "OMS テーブル定義書 - 改訂履歴"
ws0["A1"].font = Font(name="Yu Gothic", size=14, bold=True)
hist_headers = ["版", "改訂日", "改訂者", "改訂内容"]
hist_widths  = [6, 14, 14, 90]
apply_header(ws0, 3, hist_headers, hist_widths)
hist = [
    ("1.0", "2008-04-15", "田中",  "初版作成 (OMS v1.0 リリースに合わせ)"),
    ("2.0", "2011-08-02", "佐藤",  "在庫サブシステム追加に伴い T_INVENTORY 等を追記"),
    ("2.1", "2013-05-30", "佐藤",  "T_INVENTORY に VERSION 列追加 (楽観排他用)"),  # ※ 実 DDL は未追加
    ("2.2", "2014-11-30", "田中",  "M_USER.PASSWORD を SHA-256 へ移行 (※ 実装側は MD5 のまま)"),
    ("2.3", "2016-04-10", "佐藤",  "在庫移動ログ T_INVENTORY_TRX 追加"),  # ※ 実テーブルは InventoryTrx
    ("2.4", "2018-09-21", "鈴木",  "受注テーブルのインデックス IDX_ORDER_DT 追加 (※ 実 DDL に反映漏れ)"),
]
for i, row in enumerate(hist, start=4):
    write_row(ws0, i, list(row))


# ---- 各テーブル ----

make_table_sheet(
    wb, "M_USER",
    {"phys": "M_USER", "logical": "ユーザマスタ",
     "purpose": "OMS 利用者を管理する",
     "created": "2008-04-15", "updated": "2014-11-30", "author": "田中"},
    columns=[
        {"phys": "USER_ID",  "logical": "ユーザID", "type": "VARCHAR2(20)", "nn": True,  "pk": True,  "note": "主キー"},
        {"phys": "USER_NM",  "logical": "氏名",     "type": "VARCHAR2(60)", "nn": False, "note": ""},
        {"phys": "PASSWORD", "logical": "パスワード", "type": "VARCHAR2(64)", "nn": True,
         "note": "SHA-256 ハッシュ値 (16進64桁) を格納。2014 改修にて MD5 から移行"},
        {"phys": "DEPT_CD",  "logical": "所属部門コード", "type": "VARCHAR2(10)", "nn": False, "note": "M_DEPT.DEPT_CD を参照"},
        {"phys": "FAIL_CNT", "logical": "認証失敗回数", "type": "NUMBER(2)", "nn": False,
         "note": "3 回連続失敗で LOCK_FLG='1'。深夜バッチで日次クリア"},
        {"phys": "LOCK_FLG", "logical": "ロックフラグ", "type": "CHAR(1)", "nn": False, "note": "1: ロック中"},
        {"phys": "DEL_FLG",  "logical": "削除フラグ", "type": "CHAR(1)", "nn": True,  "note": "0: 有効 / 1: 削除"},
        {"phys": "CRT_DT",   "logical": "作成日時", "type": "DATE", "nn": False, "note": ""},
        {"phys": "UPD_DT",   "logical": "更新日時", "type": "DATE", "nn": False, "note": ""},
    ],
    indexes=[
        {"name": "PK_M_USER",   "kind": "PRIMARY", "cols": "USER_ID"},
        {"name": "IDX_M_USER_01", "kind": "INDEX", "cols": "DEPT_CD, DEL_FLG", "note": "部門別検索用"},
    ],
)

make_table_sheet(
    wb, "M_CUSTOMER",
    {"phys": "M_CUSTOMER", "logical": "取引先マスタ",
     "purpose": "受注先となる取引先企業を管理する",
     "created": "2008-04-15", "updated": "2016-04-10", "author": "佐藤"},
    columns=[
        {"phys": "CUSTOMER_CD", "logical": "取引先コード", "type": "VARCHAR2(10)", "nn": True, "pk": True, "note": "主キー"},
        {"phys": "CUSTOMER_NM", "logical": "取引先名",     "type": "VARCHAR2(120)","nn": True,  "note": "半角カナ・機種依存文字使用不可（運用ルール）"},
        {"phys": "TEL",         "logical": "電話番号",     "type": "VARCHAR2(20)", "nn": False, "note": ""},
        {"phys": "ADDR",        "logical": "住所",         "type": "VARCHAR2(200)","nn": False, "note": ""},
        {"phys": "DEL_FLG",     "logical": "削除フラグ",   "type": "CHAR(1)",      "nn": True,  "note": "0: 有効 / 1: 削除"},
    ],
    indexes=[
        {"name": "PK_M_CUSTOMER",   "kind": "PRIMARY", "cols": "CUSTOMER_CD"},
        {"name": "IDX_M_CUSTOMER_01","kind": "INDEX",  "cols": "CUSTOMER_NM", "note": "取引先名部分一致検索用"},
    ],
)

make_table_sheet(
    wb, "M_ITEM",
    {"phys": "M_ITEM", "logical": "商品マスタ",
     "purpose": "販売対象の商品を管理する",
     "created": "2008-04-15", "updated": "2011-08-02", "author": "佐藤"},
    columns=[
        {"phys": "ITEM_CD",     "logical": "商品コード",   "type": "VARCHAR2(13)", "nn": True, "pk": True, "note": "JAN コード 13 桁"},
        {"phys": "ITEM_NM",     "logical": "商品名",       "type": "VARCHAR2(120)","nn": True,  "note": ""},
        {"phys": "UNIT_PRICE",  "logical": "単価",         "type": "NUMBER(10,2)", "nn": True,  "note": "税抜"},
        {"phys": "CATEGORY_CD", "logical": "カテゴリコード","type": "VARCHAR2(6)",  "nn": False, "note": ""},
        {"phys": "DEL_FLG",     "logical": "削除フラグ",   "type": "CHAR(1)",      "nn": True,  "note": "0: 有効 / 1: 削除"},
    ],
    indexes=[
        {"name": "PK_M_ITEM",     "kind": "PRIMARY", "cols": "ITEM_CD"},
        {"name": "IDX_M_ITEM_01", "kind": "INDEX",   "cols": "CATEGORY_CD"},
    ],
)

make_table_sheet(
    wb, "T_ORDER",
    {"phys": "T_ORDER", "logical": "受注ヘッダ",
     "purpose": "受注 1 件分のヘッダ情報を管理する",
     "created": "2008-04-15", "updated": "2018-09-21", "author": "鈴木"},
    columns=[
        {"phys": "ORDER_NO",    "logical": "受注番号",    "type": "VARCHAR2(14)", "nn": True, "pk": True, "note": "YYYYMMDD + 6桁連番"},
        {"phys": "ORDER_DT",    "logical": "受注日",      "type": "DATE",         "nn": True,  "note": ""},
        {"phys": "CUSTOMER_CD", "logical": "取引先コード","type": "VARCHAR2(10)", "nn": True,  "note": "M_CUSTOMER.CUSTOMER_CD"},
        {"phys": "USER_ID",     "logical": "担当者ID",    "type": "VARCHAR2(20)", "nn": True,  "note": "M_USER.USER_ID"},
        {"phys": "STATUS_CD",   "logical": "ステータス",  "type": "VARCHAR2(2)",  "nn": True,
         "note": "00: 仮登録 / 10: 確定 / 90: 取消"},
        {"phys": "TOTAL_AMT",   "logical": "合計金額",    "type": "NUMBER(12,2)", "nn": True,  "note": "税込（2019 改修より）"},
        {"phys": "REMARKS",     "logical": "備考",        "type": "VARCHAR2(400)","nn": False, "note": ""},
        {"phys": "CRT_DT",      "logical": "作成日時",    "type": "DATE",         "nn": False, "note": ""},
        {"phys": "UPD_DT",      "logical": "更新日時",    "type": "DATE",         "nn": False, "note": ""},
    ],
    indexes=[
        {"name": "PK_T_ORDER",     "kind": "PRIMARY", "cols": "ORDER_NO"},
        {"name": "IDX_T_ORDER_01", "kind": "INDEX",   "cols": "ORDER_DT, STATUS_CD", "note": "一覧画面用 (2018 追加)"},
        {"name": "IDX_T_ORDER_02", "kind": "INDEX",   "cols": "CUSTOMER_CD",         "note": "取引先別検索用 (2018 追加)"},
    ],
)

make_table_sheet(
    wb, "T_ORDER_DTL",
    {"phys": "T_ORDER_DTL", "logical": "受注明細",
     "purpose": "受注 1 件分の明細行を管理する",
     "created": "2008-04-15", "updated": "2008-04-15", "author": "田中"},
    columns=[
        {"phys": "ORDER_NO",   "logical": "受注番号", "type": "VARCHAR2(14)", "nn": True, "pk": True, "note": "T_ORDER.ORDER_NO"},
        {"phys": "LINE_NO",    "logical": "明細番号", "type": "NUMBER(4)",    "nn": True, "pk": True, "note": "1〜"},
        {"phys": "ITEM_CD",    "logical": "商品コード","type": "VARCHAR2(13)","nn": True,  "note": ""},
        {"phys": "QTY",        "logical": "数量",     "type": "NUMBER(8)",   "nn": True,  "note": ""},
        {"phys": "UNIT_PRICE", "logical": "単価",     "type": "NUMBER(10,2)","nn": True,  "note": "受注確定時の単価をスナップショット"},
    ],
    indexes=[
        {"name": "PK_T_ORDER_DTL", "kind": "PRIMARY", "cols": "ORDER_NO, LINE_NO"},
    ],
)

make_table_sheet(
    wb, "T_INVENTORY",
    {"phys": "T_INVENTORY", "logical": "在庫マスタ",
     "purpose": "商品×倉庫の現在在庫数を管理する",
     "created": "2011-08-02", "updated": "2013-05-30", "author": "佐藤"},
    columns=[
        {"phys": "ITEM_CD",      "logical": "商品コード", "type": "VARCHAR2(13)", "nn": True, "pk": True, "note": ""},
        {"phys": "WAREHOUSE_CD", "logical": "倉庫コード", "type": "VARCHAR2(4)",  "nn": True, "pk": True,
         "note": "W001: 本社 / W002: 大阪 / W003: 福岡 (2022 追加)"},
        {"phys": "STOCK_QTY",    "logical": "在庫数",     "type": "NUMBER(10)",   "nn": True,  "note": ""},
        {"phys": "VERSION",      "logical": "バージョン", "type": "NUMBER(10)",   "nn": True,
         "note": "楽観排他用。UPDATE 時に WHERE VERSION = ? を必須化 (2013 改修)"},
        {"phys": "UPD_DT",       "logical": "更新日時",   "type": "DATE",         "nn": False, "note": ""},
    ],
    indexes=[
        {"name": "PK_T_INVENTORY", "kind": "PRIMARY", "cols": "ITEM_CD, WAREHOUSE_CD"},
    ],
)

make_table_sheet(
    wb, "T_INVENTORY_TRX",
    {"phys": "T_INVENTORY_TRX", "logical": "在庫移動ログ",
     "purpose": "入出庫の履歴を管理する",
     "created": "2016-04-10", "updated": "2016-04-10", "author": "佐藤"},
    columns=[
        {"phys": "TRX_ID",       "logical": "トランザクションID", "type": "NUMBER(15)",  "nn": True, "pk": True, "note": "SEQ_INV_TRX"},
        {"phys": "ITEM_CD",      "logical": "商品コード", "type": "VARCHAR2(13)", "nn": True, "note": ""},
        {"phys": "WAREHOUSE_CD", "logical": "倉庫コード", "type": "VARCHAR2(4)",  "nn": True, "note": ""},
        {"phys": "DELTA_QTY",    "logical": "増減数",     "type": "NUMBER(10)",   "nn": True, "note": "入庫: 正 / 出庫: 負"},
        {"phys": "TRX_TYPE",     "logical": "区分",       "type": "VARCHAR2(10)", "nn": True, "note": "IN / OUT / ADJ"},
        {"phys": "REF_ORDER_NO", "logical": "参照受注番号","type": "VARCHAR2(14)","nn": False,"note": ""},
        {"phys": "CREATED_AT",   "logical": "作成日時",   "type": "TIMESTAMP",    "nn": True, "note": ""},
    ],
    indexes=[
        {"name": "PK_T_INVENTORY_TRX", "kind": "PRIMARY", "cols": "TRX_ID"},
        {"name": "IDX_TRX_01", "kind": "INDEX", "cols": "ITEM_CD, CREATED_AT"},
    ],
)

make_table_sheet(
    wb, "T_INVENTORY_CACHE",
    {"phys": "T_INVENTORY_CACHE", "logical": "在庫キャッシュ",
     "purpose": "在庫照会高速化のためのキャッシュテーブル。日次バッチ rebuildAllCache() で再構築",
     "created": "2012-06-10", "updated": "2012-06-10", "author": "佐藤"},
    columns=[
        {"phys": "ITEM_CD",      "logical": "商品コード", "type": "VARCHAR2(13)", "nn": True, "pk": True, "note": ""},
        {"phys": "WAREHOUSE_CD", "logical": "倉庫コード", "type": "VARCHAR2(4)",  "nn": True, "pk": True, "note": ""},
        {"phys": "STOCK_QTY",    "logical": "在庫数",     "type": "NUMBER(10)",   "nn": True, "note": ""},
        {"phys": "CACHED_AT",    "logical": "キャッシュ生成日時", "type": "TIMESTAMP", "nn": True, "note": ""},
    ],
    indexes=[
        {"name": "PK_T_INVENTORY_CACHE", "kind": "PRIMARY", "cols": "ITEM_CD, WAREHOUSE_CD"},
    ],
)

# 改訂履歴を先頭に
wb.move_sheet("改訂履歴", offset=-len(wb.sheetnames))

wb.save(OUT)
print(f"OK: {OUT}")
