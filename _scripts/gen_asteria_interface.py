# -*- coding: utf-8 -*-
"""Asteria 連携 I/F 一覧 (Excel).

故意の不一致:
- I/F のスケジュールが flow XML の cron 設定と乖離 (例: 設計書 03:00 / 実体 02:30)
- 設計書では DWH 側 MERGE と記載するも実装は TRUNCATE
- マッピングシートで TAX_RATE の値域に「2014 改定で 0.08 標準」と記載されたまま放置
"""

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = Path(r"C:\sparticle\poc-investigation-targets\asteria-etl-sample\docs\連携IF一覧.xlsx")
OUT.parent.mkdir(parents=True, exist_ok=True)

HEADER_FILL = PatternFill("solid", fgColor="305496")
HEADER_FONT = Font(name="Yu Gothic", size=10, bold=True, color="FFFFFF")
NORMAL_FONT = Font(name="Yu Gothic", size=10)
THIN = Side(style="thin", color="999999")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)


def header(ws, row, headers, widths):
    for c, (h, w) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(c)].width = w


def row(ws, r, values, align=None):
    for c, v in enumerate(values, start=1):
        cell = ws.cell(row=r, column=c, value=v)
        cell.font = NORMAL_FONT
        cell.alignment = align or LEFT
        cell.border = BORDER


wb = Workbook()
del wb["Sheet"]

# === 連携I/F 一覧 ===
ws = wb.create_sheet("連携I_F一覧")
ws.sheet_view.showGridLines = False
ws["A1"] = "ASTERIA Warp 連携 I/F 一覧"
ws["A1"].font = Font(name="Yu Gothic", size=14, bold=True)
ws["A2"] = "最終更新: 2019-10-01 / 担当: 鈴木"
ws["A2"].font = Font(name="Yu Gothic", size=9, italic=True, color="808080")

header(
    ws, 4,
    ["I/F ID", "I/F 名", "送信元", "送信先", "プロトコル", "形式", "文字コード", "スケジュール", "更新方式", "備考"],
    [9, 26, 18, 18, 10, 10, 11, 16, 14, 30],
)
ifs = [
    ("AIF-001", "取引先マスタ連携",      "OMS (Oracle)",       "DWH (SQL Server)", "JDBC",        "テーブル", "UTF-8", "平日 06:30", "MERGE",          "FLW-001"),
    ("AIF-002", "受注ヘッダ連携",        "OMS (Oracle)",       "DWH (SQL Server)", "JDBC + ファイル", "TSV",   "UTF-8", "毎日 03:00", "MERGE (差分)",   "FLW-002 / 中間 TSV 経由"),
    ("AIF-003", "在庫マスタ突合",        "OMS (Oracle) / DWH", "ファイル共有",      "JDBC + ファイル", "CSV",   "Shift_JIS", "月初 04:30", "レポート出力のみ", "FLW-003"),
    ("AIF-099", "旧 BI 向け CSV 出力",    "OMS (Oracle)",       "ファイル共有",      "JDBC + ファイル", "CSV",   "Shift_JIS", "手動",       "全件",           "FLW-099 / 廃止予定"),
]
for i, f in enumerate(ifs, start=5):
    row(ws, i, list(f))

# === AIF-002 マッピング ===
ws = wb.create_sheet("AIF-002マッピング")
ws.sheet_view.showGridLines = False
ws["A1"] = "AIF-002 受注ヘッダ連携 - マッピング定義書"
ws["A1"].font = Font(name="Yu Gothic", size=14, bold=True)

ws["A3"] = "■ 基本情報"
ws["A3"].font = Font(name="Yu Gothic", size=11, bold=True)
meta = [
    ("送信元",       "OMS Oracle T_ORDER (+ M_CUSTOMER LookupDB)"),
    ("送信先",       "DWH SQL Server FACT_ORDER"),
    ("実行頻度",     "毎日 03:00 (前日分のみを取得する差分連携)"),
    ("コンポーネント数", "6 (O1〜O6)"),
    ("関連ファイル", "flows/order_etl_daily.flwx / mappers/order_transform.xsl"),
]
for i, (k, v) in enumerate(meta, start=4):
    ws.cell(row=i, column=1, value=k).font = Font(name="Yu Gothic", size=10, bold=True)
    ws.cell(row=i, column=1).fill = PatternFill("solid", fgColor="DDEBF7")
    ws.cell(row=i, column=1).border = BORDER
    ws.cell(row=i, column=1).alignment = LEFT
    c2 = ws.cell(row=i, column=2, value=v)
    c2.font = NORMAL_FONT
    c2.border = BORDER
    c2.alignment = LEFT
    ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=7)
ws.column_dimensions["A"].width = 18
for col in "BCDEFG":
    ws.column_dimensions[col].width = 16

r = 4 + len(meta) + 1
ws.cell(row=r, column=1, value="■ 項目マッピング").font = Font(name="Yu Gothic", size=11, bold=True)
r += 1
header(
    ws, r,
    ["No", "送信元項目", "型", "変換ロジック", "送信先項目", "型", "備考"],
    [4, 22, 14, 36, 22, 14, 28],
)
r += 1
mappings = [
    (1,  "T_ORDER.ORDER_NO",     "VARCHAR2(14)",  "(無変換)",                                  "FACT_ORDER.ORDER_NO",     "VARCHAR(14)",   ""),
    (2,  "T_ORDER.ORDER_DT",     "DATE",          "(無変換)",                                  "FACT_ORDER.ORDER_DT",     "DATE",          ""),
    (3,  "T_ORDER.CUSTOMER_CD",  "VARCHAR2(10)",  "(無変換)",                                  "FACT_ORDER.CUSTOMER_CD",  "VARCHAR(10)",   ""),
    (4,  "LookupDB CUSTOMER_NM", "VARCHAR2(120)", "M_CUSTOMER から Lookup",                    "FACT_ORDER.CUSTOMER_NM",  "NVARCHAR(120)", "LookupDB cache=true 推奨"),
    (5,  "T_ORDER.STATUS_CD",    "VARCHAR2(2)",   "(無変換)",                                  "FACT_ORDER.STATUS_CD",    "VARCHAR(2)",    ""),
    (6,  "T_ORDER.TOTAL_AMT",    "NUMBER(12,2)",  "TOTAL_AMT を税抜換算 (1 + TAX_RATE で除算)", "FACT_ORDER.TAX_EXCL_AMT", "DECIMAL(14,2)", "XSLT で計算"),
    (7,  "T_ORDER.REMARKS",      "VARCHAR2(400)", "備考欄から '軽減' を検索し税率を判定。標準 0.08", "FACT_ORDER.TAX_RATE",     "DECIMAL(5,3)",  "2014 改定対応"),
    (8,  "T_ORDER.TOTAL_AMT",    "NUMBER(12,2)",  "TAX_EXCL_AMT × TAX_RATE で再計算後 + TAX_EXCL_AMT", "FACT_ORDER.TOTAL_AMT",    "DECIMAL(14,2)", ""),
    (9,  "(固定値)",             "-",             "'STORE' 固定",                              "FACT_ORDER.SALES_CHANNEL","VARCHAR(10)",   ""),
    (10, "(固定値)",             "-",             "'JPY' 固定",                                "FACT_ORDER.CURRENCY",     "VARCHAR(3)",    ""),
    (11, "SYSDATE",              "-",             "ETL 実行時のシステム日時",                  "FACT_ORDER.LOAD_AT",      "DATETIME2",     ""),
]
for m in mappings:
    row(ws, r, list(m))
    r += 1

r += 1
ws.cell(row=r, column=1, value="■ 異常系").font = Font(name="Yu Gothic", size=11, bold=True)
r += 1
header(ws, r, ["No", "ケース", "対応"], [6, 30, 60])
r += 1
errors = [
    ("E-01", "OMS 接続失敗",   "Warp 内蔵リトライで 3 回 (5/15/30 分間隔)。失敗で it-ops 通知。"),
    ("E-02", "DWH BULK INSERT 失敗", "DWH 側で取込前にステージング検証。失敗時はロールバックして翌実行に持ち越し。"),
    ("E-03", "LookupDB ヒット無し",  "DWH の CUSTOMER_NM は NULL を許容しないため、ETL 側で取引先名を '(未登録)' で代替。"),
]
for e in errors:
    row(ws, r, list(e))
    r += 1

wb.save(OUT)
print("OK")
